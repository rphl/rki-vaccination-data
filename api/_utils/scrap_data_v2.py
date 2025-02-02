""" Scrap Data from Excel sheet """
import re
import datetime
from io import BytesIO
import urllib.request
import openpyxl
from .statics import inhabitants


def get_file():
  """ Get remote file content """
  # Request to load excel sheet
  url = 'https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten' \
      '/Impfquotenmonitoring.xlsx?__blob=publicationFile'
  hdr = {
      'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11'
      ' (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11'}

  req = urllib.request.Request(url, headers=hdr)
  with urllib.request.urlopen(req) as response:
    return response.read()


def get_data():
  """ Get Data for API """
  file = get_file()

  # Read excel sheet
  work_book = openpyxl.load_workbook(filename=BytesIO(file))
  sheet = work_book[work_book.sheetnames[2]]

  # Load update time
  first_sheet = work_book[work_book.sheetnames[0]]
  relast_update_match = re.search(
      r"[\d]{2}\.[\d]{2}\.[\d]{2}", first_sheet['A3'].value)
  last_update = datetime.datetime.strptime(
      relast_update_match.group(), '%d.%m.%y')

  # Load data from rows
  rows = []
  for row in sheet.iter_rows(max_row=21):
    if row[1].value is None:
      continue
    state = row[1].value.replace("*", "").strip()
    if state in inhabitants.STATES or state in ['Bundesressorts', 'Gesamt']:
      if state == 'Bundesressorts':
        total = 0
      elif state == 'Gesamt':
        total = inhabitants.TOTAL
        state = "Deutschland"
      else:
        total = inhabitants.STATES[state]['total']
      data = {
        "name": state,
        "inhabitants": total,
        "isState": state in inhabitants.STATES,
        "rs": str(row[0].value) if row[0].value is not None else "",
        "vaccinatedAtLeastOnce": {
          "doses": row[2].value,
          "quote": round(row[2].value / total * 100, 2) if total != 0 else 0,
          "differenceToThePreviousDay": row[7].value,
          "vaccine": [
            {
              "name": "biontech",
              "doses": row[3].value
            },
            {
              "name": "moderna",
              "doses": row[4].value
            },
            {
              "name": "astrazeneca",
              "doses": row[5].value
            },
            {
              "name": "janssen",
              "doses": row[6].value
            }
          ]
        },
        "fullyVaccinated": {
          "doses": row[8].value + row[6].value,
          "quote": round((row[8].value + row[6].value) / total * 100, 2) if total != 0 else 0,
          "differenceToThePreviousDay": row[12].value,
          "vaccine": [
            {
              "name": "biontech",
              "firstDoses": row[3].value,
              "secondDoses": row[9].value,
              "totalDoses": row[3].value + row[9].value
            },
            {
              "name": "moderna",
              "firstDoses": row[4].value,
              "secondDoses": row[10].value,
              "totalDoses": row[4].value + row[10].value
            },
            {
              "name": "astrazeneca",
              "firstDoses": row[5].value,
              "secondDoses": row[11].value,
              "totalDoses": row[5].value + row[11].value
            },
            {
              "name": "janssen",
              "firstDoses": row[6].value,
              "totalDoses": row[6].value
            }
          ]
        },
        "boosterVaccinated": {
          "doses": row[13].value,
          "quote": round(row[13].value / total * 100, 2) if total != 0 else 0,
          "differenceToThePreviousDay": row[17].value,
          "vaccine": [
            {
              "name": "biontech",
              "firstDoses": row[3].value,
              "secondDoses": row[9].value,
              "boosterDoses": row[14].value,
              "totalDoses": row[3].value + row[9].value + row[14].value
            },
            {
              "name": "moderna",
              "firstDoses": row[4].value,
              "secondDoses": row[10].value,
              "boosterDoses": row[15].value,
              "totalDoses": row[4].value + row[10].value + row[15].value
            },
            {
              "name": "janssen",
              "firstDoses": row[6].value,
              "boosterDoses": row[16].value,
              "totalDoses": row[6].value + row[16].value
            }
          ]
        }        
      }
      rows.append(data)
  return {
    "lastUpdate": last_update,
    "data": rows
  }
